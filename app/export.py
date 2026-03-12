"""
报告导出模块
支持 CSV、Excel 格式导出
"""

import os
import csv
import json
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
from loguru import logger

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl 未安装，Excel 导出功能不可用")


class ReportExporter:
    """报告导出器"""
    
    def __init__(self, output_dir: str = "data/exports"):
        """
        初始化导出器
        
        Args:
            output_dir: 输出目录
        """
        self.output_dir = output_dir
        self._ensure_directory()
        logger.info(f"报告导出器已初始化：{output_dir}")
    
    def _ensure_directory(self):
        """确保输出目录存在"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"创建导出目录：{self.output_dir}")
    
    def _generate_filename(self, prefix: str, extension: str) -> str:
        """生成带时间戳的文件名"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return os.path.join(self.output_dir, f"{prefix}_{timestamp}.{extension}")
    
    def export_to_csv(
        self, 
        data: List[Dict[str, Any]], 
        filename: Optional[str] = None
    ) -> str:
        """
        导出为 CSV
        
        Args:
            data: 数据列表
            filename: 文件名（可选）
        
        Returns:
            输出文件路径
        """
        if filename is None:
            filename = self._generate_filename('missing_episodes', 'csv')
        
        if not data:
            logger.warning("没有数据可导出")
            return filename
        
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            # 获取所有字段名
            fieldnames = list(data[0].keys())
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            
            writer.writeheader()
            for row in data:
                # 处理列表字段
                processed_row = {}
                for key, value in row.items():
                    if isinstance(value, list):
                        processed_row[key] = ', '.join(map(str, value))
                    elif isinstance(value, datetime):
                        processed_row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        processed_row[key] = value
                writer.writerow(processed_row)
        
        logger.info(f"已导出 {len(data)} 条记录到 CSV: {filename}")
        return filename
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: str = "缺集报告"
    ) -> Optional[str]:
        """
        导出为 Excel
        
        Args:
            data: 数据列表
            filename: 文件名（可选）
            sheet_name: 工作表名称
        
        Returns:
            输出文件路径，如果 openpyxl 不可用则返回 None
        """
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl 未安装，无法导出 Excel")
            return None
        
        if filename is None:
            filename = self._generate_filename('missing_episodes', 'xlsx')
        
        if not data:
            logger.warning("没有数据可导出")
            return filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 样式定义
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                
                # 处理特殊类型
                if isinstance(value, list):
                    value = ', '.join(map(str, value))
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 添加统计信息工作表
        self._add_summary_sheet(wb, data)
        
        wb.save(filename)
        logger.info(f"已导出 {len(data)} 条记录到 Excel: {filename}")
        return filename
    
    def _add_summary_sheet(self, wb: Workbook, data: List[Dict]):
        """添加统计信息工作表"""
        ws = wb.create_sheet("统计摘要")
        
        # 统计数据
        total_series = len(set(row.get('series_name', '') for row in data))
        total_missing = sum(
            len(row.get('episode_numbers', [])) if isinstance(row.get('episode_numbers'), list) 
            else 0 for row in data
        )
        
        summary_data = [
            ("报告生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("剧集总数", str(total_series)),
            ("缺失总集数", str(total_missing)),
            ("记录总数", str(len(data)))
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
    
    def export_detection_result(self, result) -> str:
        """
        导出检测结果为 Excel
        
        Args:
            result: DetectionResult 对象
        
        Returns:
            输出文件路径
        """
        from app.detector import MissingEpisodeDetector
        
        # 转换为可导出格式
        data = []
        for series in result.series:
            if series.missing_episodes_count > 0:
                for season in series.seasons:
                    if season.missing_episodes:
                        data.append({
                            'series_name': series.series_name,
                            'series_id': series.series_id,
                            'season': season.season_number,
                            'episode_numbers': season.missing_episodes,
                            'missing_count': len(season.missing_episodes)
                        })
        
        filename = self._generate_filename('emby_missing_report', 'xlsx')
        self.export_to_excel(data, filename, "缺集列表")
        
        return filename


def setup_export_routes(app, detector, db):
    """为 FastAPI 应用添加导出路由"""
    from fastapi import Response
    from fastapi.responses import FileResponse
    
    exporter = ReportExporter()
    
    @app.get("/api/export/csv")
    async def export_csv():
        """导出 CSV"""
        if db is None:
            return {"error": "数据库未初始化"}
        
        data = db.get_latest_missing_episodes()
        filename = exporter.export_to_csv(data)
        
        return FileResponse(
            filename,
            media_type='text/csv',
            filename=os.path.basename(filename)
        )
    
    @app.get("/api/export/excel")
    async def export_excel():
        """导出 Excel"""
        if db is None:
            return {"error": "数据库未初始化"}
        
        data = db.get_latest_missing_episodes()
        filename = exporter.export_to_excel(data)
        
        if filename:
            return FileResponse(
                filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=os.path.basename(filename)
            )
        else:
            return {"error": "Excel 导出不可用，请安装 openpyxl"}
